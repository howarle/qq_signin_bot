import os
import openpyxl
import datetime

from SignIn.userconfig import *

recordsheet_file = os.path.split(os.path.realpath(__file__))[
    0] + '/signin.xlsx'


def is_signin(name: str):
    return (sheet_getdata(name) != None) and (sheet_getdata(name) != '咕')

def is_submit_signed(name: str):
    return (sheet_getdata(name) != None)

def sign_it(name: str):
    current_time = datetime.datetime.strftime(datetime.datetime.now(), '%H:%M')
    return sheet_change(name, current_time)

def sign_gugugu(name: str):
    return sheet_change(name, '咕')

def revoke_sign(name: str):
    return sheet_change(name, None)


def sheet_init():
    if os.path.exists(recordsheet_file):
        return
    book = openpyxl.Workbook()
    sh = book.active
    sh.title = '1'

    sh.cell(1, 1).value = '姓名'
    sh.cell(1, 2).value = 'Q号'
    cnt = 1
    for user in qq_to_name:
        cnt = cnt + 1
        sh.cell(cnt, 1).value = qq_to_name[user]
        sh.cell(cnt, 2).value = user

    book.save(recordsheet_file)


def sheet_change(name: str, content):
    sheet_init()

    book = openpyxl.load_workbook(recordsheet_file)
    sh = book['1']

    row = 2
    while sh.cell(row, 1).value != None and sh.cell(row, 1).value != name:
        row = row + 1

    if sh.cell(row, 1).value != name:
        return False

    today_str = str(datetime.date.today())
    col = 3
    while sh.cell(1, col).value != None and sh.cell(1, col).value != today_str:
        col = col + 1

    if sh.cell(1, col).value != today_str:
        sh.cell(1, col).value = today_str

    sh.cell(row, col).value = content

    book.save(recordsheet_file)
    return True


def sheet_getdata(name: str):
    sheet_init()

    book = openpyxl.load_workbook(recordsheet_file)
    sh = book['1']

    row = 2
    while sh.cell(row, 1).value != None and sh.cell(row, 1).value != name:
        row = row + 1

    if sh.cell(row, 1).value != name:
        return None

    today_str = str(datetime.date.today())
    col = 3
    while sh.cell(1, col).value != None and sh.cell(1, col).value != today_str:
        col = col + 1

    if sh.cell(1, col).value != today_str:
        return None

    return sh.cell(row, col).value
